# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# MIT License. See license.txt
from __future__ import unicode_literals

import frappe

import openpyxl
import xlrd
import re
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
from six import BytesIO, string_types

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
# return xlsx file object
def make_xlsx(data, sheet_name, wb=None):

	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)

	thin_border = Border(left=Side(style='thin',  color="FF000000"),
                     right=Side(style='thin',  color="FF000000"),
                     top=Side(style='thin',  color="FF000000"),
                     bottom=Side(style='thin', color = "FF000000"))

	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, string_types) and (sheet_name not in ['Data Import Template', 'Data Export']):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, string_types) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = re.sub(ILLEGAL_CHARACTERS_RE, '', value)

			clean_row.append(value)

		ws.append(clean_row)

	#changing font and borders for the output cells
	header_flag = True
	for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1):
		no_of_values = sum([1 if cell.value!=None else 0 for cell in row])
		for cell in row:

			#change header to bold if there is only 1 value in the whole row - usually header items or of the header row
			if (no_of_values == 1) or (no_of_values > 1 and header_flag):
				cell.font = Font(bold=True)

			#add borders to all cells in the table
			if no_of_values > 1 and cell.value != None:
				cell.border = thin_border

		#once we've iterated over the header row, change flag
		if no_of_values > 1:
			header_flag = False

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file


def handle_html(data):
	# return if no html tags found
	data = frappe.as_unicode(data)

	if '<' not in data:
		return data
	if '>' not in data:
		return data

	from html2text import HTML2Text

	h = HTML2Text()
	h.unicode_snob = True
	h = h.unescape(data or "")

	obj = HTML2Text()
	obj.ignore_links = True
	obj.body_width = 0

	try:
		value = obj.handle(h)
	except Exception:
		# unable to parse html, send it raw
		return data

	value = ", ".join(value.split('  \n'))
	value = " ".join(value.split('\n'))
	value = ", ".join(value.split('# '))

	return value

def read_xlsx_file_from_attached_file(file_url=None, fcontent=None, filepath=None):
	if file_url:
		_file = frappe.get_doc("File", {"file_url": file_url})
		filename = _file.get_full_path()
	elif fcontent:
		from io import BytesIO
		filename = BytesIO(fcontent)
	elif filepath:
		filename = filepath
	else:
		return

	rows = []
	wb1 = load_workbook(filename=filename, read_only=True, data_only=True)
	ws1 = wb1.active
	for row in ws1.iter_rows():
		tmp_list = []
		for cell in row:
			tmp_list.append(cell.value)
		rows.append(tmp_list)
	return rows

def read_xls_file_from_attached_file(content):
	book = xlrd.open_workbook(file_contents=content)
	sheets = book.sheets()
	sheet = sheets[0]
	rows = []
	for i in range(sheet.nrows):
		rows.append(sheet.row_values(i))
	return rows

def build_xlsx_response(data, filename):
	xlsx_file = make_xlsx(data, filename)
	# write out response as a xlsx type
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'
